import streamlit as st
import io
import re
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.title import Title
from openpyxl.chart.text import RichText, Text
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties,
    Font as DrawingFont, RegularTextRun
)
from openpyxl.chart.label import DataLabelList  # 추가됨
from openpyxl.chart.axis import ChartLines      # 추가됨

# =====================================================================
# 1. 엑셀 스타일 및 유틸 함수 모음
# =====================================================================
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=9)
BLK_FILL = PatternFill("solid", fgColor="D9E1F2")
BLK_FONT = Font(name=FONT, bold=True, size=9)
LBL_FILL = PatternFill("solid", fgColor="F2F2F2")
BASE_FONT = Font(name=FONT, size=9)
AVG_FILL = PatternFill("solid", fgColor="FCE4D6")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
INPUT_FONT = Font(name=FONT, size=9, color="0000FF")
TITLE_FONT = Font(name=FONT, bold=True, size=13)
NOTE_FONT = Font(name=FONT, size=8, italic=True, color="808080")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _find_col_by_priority(ws, candidates, scan_rows=6, scan_cols=100):
    for kw in candidates:
        for r in range(1, scan_rows + 1):
            for c in range(1, scan_cols + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and kw in v.replace(" ", ""):
                    return c
    return None

def detect_sewage_columns(ws, override=None):
    if override: return override
    date_col = _find_col_by_priority(ws, ["날짜"])
    inflow_col = _find_col_by_priority(ws, ["실제유입량", "유입량(반류수포함)", "유입량"])
    bod_col = _find_col_by_priority(ws, ["BOD"])
    toc_col = _find_col_by_priority(ws, ["TOC"])
    is_cod = False
    if toc_col is None:
        toc_col = _find_col_by_priority(ws, ["COD"])
        is_cod = toc_col is not None
    ss_col = _find_col_by_priority(ws, ["SS"])
    tn_col = _find_col_by_priority(ws, ["T-N", "TN"])
    tp_col = _find_col_by_priority(ws, ["T-P", "TP"])
    return dict(date=date_col, inflow=inflow_col, BOD=bod_col, TOC=toc_col,
                SS=ss_col, TN=tn_col, TP=tp_col, is_cod=is_cod)

def extract_sewage_dataframe(uploaded_file, override=None, toc_cod_factor=0.625):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    cols = detect_sewage_columns(ws, override)
    required = ["date", "inflow", "BOD", "TOC", "SS", "TN", "TP"]
    missing = [k for k in required if cols.get(k) is None]
    if missing:
        raise ValueError(f"[{uploaded_file.name}] 다음 항목의 열을 찾지 못했습니다: {missing}")
    rows = []
    for r in range(1, ws.max_row + 1):
        dv = ws.cell(row=r, column=cols["date"]).value
        if not isinstance(dv, datetime.datetime): continue
        rows.append({
            "날짜": dv,
            "유입량": ws.cell(row=r, column=cols["inflow"]).value,
            "BOD": ws.cell(row=r, column=cols["BOD"]).value,
            "TOC": ws.cell(row=r, column=cols["TOC"]).value,
            "SS": ws.cell(row=r, column=cols["SS"]).value,
            "TN": ws.cell(row=r, column=cols["TN"]).value,
            "TP": ws.cell(row=r, column=cols["TP"]).value,
        })
    if not rows: raise ValueError(f"[{uploaded_file.name}] 날짜 데이터를 찾지 못했습니다.")
    df = pd.DataFrame(rows)
    for c in ["유입량", "BOD", "TOC", "SS", "TN", "TP"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    if cols.get("is_cod"):
        df["TOC"] = df["TOC"] * toc_cod_factor
        df.attrs["toc_converted"] = True
    else:
        df.attrs["toc_converted"] = False
    
    uploaded_file.seek(0) 
    return df, openpyxl.load_workbook(uploaded_file, data_only=True)

def extract_rainfall_dataframe(uploaded_file):
    import io
    ext = uploaded_file.name.split('.')[-1].lower()
    try:
        engine = 'openpyxl' if ext == 'xlsx' else 'xlrd'
        preview = pd.read_excel(uploaded_file, header=None, nrows=20, engine=engine)
        hdr_row = None
        for i in range(len(preview)):
            if preview.iloc[i].astype(str).str.contains("강수량").any():
                hdr_row = i
                break
        if hdr_row is None: raise ValueError("강수량 헤더 없음")
        uploaded_file.seek(0)
        raw = pd.read_excel(uploaded_file, header=hdr_row, engine=engine)
    except Exception:
        uploaded_file.seek(0)
        raw_bytes = uploaded_file.read()
        text_lines = None
        for enc in ["cp949", "utf-8", "utf-8-sig", "euc-kr"]:
            try:
                text_lines = raw_bytes.decode(enc).splitlines()
                break
            except UnicodeDecodeError: continue
        if text_lines and any("지점번호" in l for l in text_lines[:20]):
            hidx = next(i for i, l in enumerate(text_lines) if "지점번호" in l)
            raw = pd.read_csv(io.StringIO("\n".join(text_lines[hidx:])), sep="\t")
        else:
            uploaded_file.seek(0)
            raw = pd.read_csv(uploaded_file, encoding='cp949')

    date_col = next((c for c in raw.columns if "일시" in str(c) or "날짜" in str(c)), None)
    rain_col = next((c for c in raw.columns if "강수량" in str(c) and "1시간" not in str(c)), None)
    if date_col is None or rain_col is None:
        raise ValueError(f"[{uploaded_file.name}] 날짜/강수량 컬럼을 찾지 못했습니다.")

    df = raw[[date_col, rain_col]].copy()
    df.columns = ["날짜", "강수량"]
    df = df.dropna(subset=["날짜"])
    df["날짜"] = pd.to_datetime(df["날짜"])
    df["강수량"] = pd.to_numeric(df["강수량"], errors="coerce").fillna(0.0)
    uploaded_file.seek(0)
    return df.sort_values("날짜").reset_index(drop=True), df

def season_of_month(m):
    if m in (3, 4, 5): return "봄"
    if m in (6, 7, 8): return "여름"
    if m in (9, 10, 11): return "가을"
    return "겨울"

def classify_rain_event(df, event_mm, influence_days):
    rain = df["강수량"].to_numpy()
    n = len(rain)
    is_event = rain >= event_mm
    cls = []
    for i in range(n):
        if is_event[i]:
            cls.append("강우시")
        else:
            lo = max(0, i - influence_days)
            if rain[i] == 0 and is_event[lo:i].any():
                cls.append("강우영향일")
            else:
                cls.append("청천시")
    df["강우구분"] = cls
    return df

def build_master(sewage_dfs, rain_df, cfg):
    df = pd.concat(sewage_dfs, ignore_index=True).sort_values("날짜").reset_index(drop=True)
    df = df.merge(rain_df, on="날짜", how="left")
    df["강수량"] = df["강수량"].fillna(0.0)
    df["연도"] = df["날짜"].dt.year
    df["월"] = df["날짜"].dt.month
    df["계절"] = df["월"].apply(season_of_month)
    df = classify_rain_event(df, cfg["rain_event_mm"], cfg["rain_influence_days"])
    return df

def style_row(ws, row, ncol, header=False, block=False, avg_col=None, label_col=True):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        if header:
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif block:
            cell.font = BLK_FONT; cell.fill = BLK_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.font = BASE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if label_col and c == 1:
                cell.fill = LBL_FILL; cell.alignment = Alignment(horizontal="left", vertical="center")
            if avg_col and c == avg_col:
                cell.fill = AVG_FILL

def set_title(ws, row, text, ncol, subtitle=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncol - 1, 1))
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    if subtitle:
        c = ws.cell(row=row, column=ncol, value=subtitle)
        c.font = Font(name=FONT, size=9, italic=True, color="595959")
        c.alignment = Alignment(horizontal="right")

def note(ws, row, ncol, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    c.alignment = Alignment(horizontal="left", wrap_text=True)

# =====================================================================
# 차트 리치텍스트 서식 유틸 (제목/축/범례 폰트를 최종본과 동일하게)
# =====================================================================
CHART_FONT = "Calibri"
AXIS_LINE_COLOR = "B3B3B3"

def _char_props(size=1000, bold=False, color="000000", font=CHART_FONT):
    return CharacterProperties(
        sz=size, b=bold, spc=-1, strike="noStrike",
        solidFill=color, latin=DrawingFont(typeface=font)
    )

def _rich_title(text, size=1400, bold=True, color="000000", font=CHART_FONT):
    """제목/축제목용 Title 객체 (실제 텍스트 포함)"""
    cp = _char_props(size, bold, color, font)
    pp = ParagraphProperties(defRPr=cp)
    run = RegularTextRun(rPr=cp, t=text)
    para = Paragraph(pPr=pp, r=[run])
    rich = RichText(p=[para])
    t = Title(tx=Text(rich=rich))
    t.overlay = False
    t.spPr = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
    return t

def _rich_txpr(size=1000, bold=False, color="000000", font=CHART_FONT):
    """축 눈금라벨/범례 텍스트 서식용 RichText (텍스트 없이 기본 서식만)"""
    cp = _char_props(size, bold, color, font)
    pp = ParagraphProperties(defRPr=cp)
    return RichText(p=[Paragraph(pPr=pp)])

def _axis_line_props(color=AXIS_LINE_COLOR, w=9525):
    return GraphicalProperties(ln=LineProperties(solidFill=color, w=w))

def style_axis(axis, title_text=None, title_size=1000, tick_size=1000,
               number_format=None, show_gridlines=False, show_minor_gridlines=False):
    """축(가로/세로) 서식을 지정합니다. 기본 눈금 및 보조 눈금선 설정이 추가되었습니다."""
    if title_text:
        axis.title = _rich_title(title_text, size=title_size, bold=True)
    
    # 기본 가로, 기본 세로 축 적용 (눈금 표시)
    axis.majorTickMark = "out"
    axis.minorTickMark = "none"
    axis.spPr = _axis_line_props()
    axis.txPr = _rich_txpr(size=tick_size, bold=False)
    
    if number_format:
        axis.number_format = number_format
        
    if not show_gridlines:
        axis.majorGridlines = None
        
    # 기본보조가로(Minor Gridlines) 눈금선 적용
    if show_minor_gridlines:
        axis.minorGridlines = ChartLines()

# 차트를 예쁘게 꾸며주는 유틸 함수 (최종 결과물과 동일한 서식: 제목/축/범례 폰트,
# 옅은 테두리, 그리드라인 제거, 둥근 모서리 없음)
def apply_beautiful_chart_style(chart, title_text=None, title_size=1400):
    # 차트 전체 테두리/배경
    chart.graphical_properties = GraphicalProperties(
        ln=LineProperties(solidFill=ColorChoice(srgbClr="D9D9D9"), round=True, w=9360)
    )
    chart.graphical_properties.solidFill = "FFFFFF"
    try:
        chart.roundedCorners = False
    except Exception:
        pass

    # 제목 서식 (기존 문자열 제목이 있으면 리치텍스트로 치환)
    if title_text is not None:
        chart.title = _rich_title(title_text, size=title_size, bold=True)
    elif chart.title is not None:
        existing = None
        try:
            existing = chart.title.tx.rich.p[0].r[0].t
        except Exception:
            existing = None
        if existing:
            chart.title = _rich_title(existing, size=title_size, bold=True)

    # 범례 위치를 아래쪽(bottom)으로 변경
    chart.legend.position = "b"
    chart.legend.overlay = False
    chart.legend.spPr = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
    chart.legend.txPr = _rich_txpr(size=1000, bold=False)

    # x축(가로/카테고리축) 서식
    if getattr(chart, "x_axis", None) is not None:
        style_axis(chart.x_axis, tick_size=1000, show_gridlines=False)
    # y축(세로/값축) 서식: 가로 주 눈금선 및 보조 눈금선 모두 켜기
    if getattr(chart, "y_axis", None) is not None:
        style_axis(chart.y_axis, tick_size=1000, show_gridlines=True, show_minor_gridlines=True)


# =====================================================================
# 핵심: 워크북 생성 (분석 표, 차트 모두 포함)
# =====================================================================
def build_workbook(master_df, sewage_raw_wbs, rain_raw, cfg, file_names, toc_converted_flags):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    years = sorted(master_df["연도"].unique().tolist())
    n_days = len(master_df)

    # ---------------- 0_안내 ----------------
    ws = wb.create_sheet("0_안내")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{cfg['facility_name']} 유입특성 분석 - 분석조건 및 가정"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    def sec(row, text):
        ws.cell(row=row, column=1, value=text).font = Font(name=FONT, bold=True, size=10, color="1F4E78")

    def kv(row, k, v, note_="", input_cell=False):
        ws.cell(row=row, column=1, value=k).font = BASE_FONT
        ws.cell(row=row, column=1).fill = LBL_FILL
        c = ws.cell(row=row, column=3, value=v)
        c.font = INPUT_FONT if input_cell else BASE_FONT
        if input_cell:
            c.fill = INPUT_FILL
        ws.cell(row=row, column=4, value=note_).font = NOTE_FONT

    sec(3, "1. 처리시설 기본정보")
    kv(5, "시설용량(㎥/일)", cfg["capacity"], "필요시 직접 수정", input_cell=True)
    kv(6, "처리대상 시설", cfg["facility_name"], "")
    kv(7, "분석기간", f"{master_df['날짜'].min().date()} ~ {master_df['날짜'].max().date()} ({n_days}일)", "")

    sec(9, "2. 유입하수량 산정 기준")
    kv(10, "기준 컬럼", "실제유입량(반류수 제외, ㎥/일)", "")

    sec(12, "3. 강우시/청천시/강우영향일 구분 기준 (표2·표4 적용)")
    kv(13, "강우시", f"당일 강수량 ≥ {cfg['rain_event_mm']}mm", "")
    kv(14, "강우영향일", f"당일 강수량 = 0mm 이고, 직전 {cfg['rain_influence_days']}일 이내 강우시가 있었던 날", "")
    kv(15, "청천시", "위 두 경우를 제외한 나머지 전체", "")

    sec(17, "4. 강우량 구간별 기준 (표3 적용)")
    r = 18
    kv(r, "청천시(구간최소값 미만)", f"당일 강수량 < {cfg['rain_bands'][0]}mm", ""); r += 1
    for b in cfg["rain_bands"]:
        kv(r, f"강우시 {b}mm 이상", f"당일 강수량 ≥ {b}mm", "누적(포함) 기준"); r += 1

    sec(r + 1, "5. 설계 유입수질 기준값 (표5 비율 계산용)")
    r += 2
    ws.cell(row=r, column=1, value="항목").font = HDR_FONT
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=3, value="설계기준(㎎/L)").font = HDR_FONT
    ws.cell(row=r, column=3).fill = HDR_FILL
    design_cell_row = {}
    r += 1
    for item in ["BOD", "TOC", "SS", "TN", "TP"]:
        label = "T-N" if item == "TN" else ("T-P" if item == "TP" else item)
        kv(r, f"{label} (㎎/L)", cfg["design_quality"].get(item), "미입력 시 비율(%) 계산은 공란 처리", input_cell=True)
        design_cell_row[item] = r
        r += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 34

    CAP = "'0_안내'!$C$5"
    DES = {k: f"'0_안내'!$C${design_cell_row[k]}" for k in design_cell_row}

    # ---------------- 일별통합데이터 ----------------
    ws_i = wb.create_sheet("일별통합데이터")
    headers = ["날짜", "연도", "월", "계절", "강수량(mm)", "유입량(㎥/일)\n(실제유입량기준)",
               "강우구분", "BOD(㎎/L)", "TOC(㎎/L)", "SS(㎎/L)", "T-N(㎎/L)", "T-P(㎎/L)", "시설용량(상수)"]
    for i, h in enumerate(headers, start=1):
        c = ws_i.cell(row=1, column=i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws_i.freeze_panes = "A2"

    for idx, rec in enumerate(master_df.itertuples(index=False), start=2):
        ws_i.cell(row=idx, column=1, value=rec.날짜).number_format = "yyyy-mm-dd"
        ws_i.cell(row=idx, column=2, value=rec.연도)
        ws_i.cell(row=idx, column=3, value=rec.월)
        ws_i.cell(row=idx, column=4, value=rec.계절)
        ws_i.cell(row=idx, column=5, value=round(float(rec.강수량), 2))
        ws_i.cell(row=idx, column=6, value=round(float(rec.유입량), 2) if pd.notna(rec.유입량) else None)
        ws_i.cell(row=idx, column=7, value=rec.강우구분)
        ws_i.cell(row=idx, column=8, value=round(float(rec.BOD), 2) if pd.notna(rec.BOD) else None)
        ws_i.cell(row=idx, column=9, value=round(float(rec.TOC), 2) if pd.notna(rec.TOC) else None)
        ws_i.cell(row=idx, column=10, value=round(float(rec.SS), 2) if pd.notna(rec.SS) else None)
        ws_i.cell(row=idx, column=11, value=round(float(rec.TN), 2) if pd.notna(rec.TN) else None)
        ws_i.cell(row=idx, column=12, value=round(float(rec.TP), 2) if pd.notna(rec.TP) else None)
        ws_i.cell(row=idx, column=13, value=f"={CAP}")
        for c in range(1, 14):
            ws_i.cell(row=idx, column=c).font = BASE_FONT
    for c, w in zip(range(1, 14), [12, 7, 6, 8, 10, 14, 11, 10, 10, 10, 10, 10, 12]):
        ws_i.column_dimensions[get_column_letter(c)].width = w

    LASTROW = n_days + 1
    INT = "일별통합데이터"
    Bc = f"{INT}!$B$2:$B${LASTROW}"
    Fc = f"{INT}!$F$2:$F${LASTROW}"
    Gc = f"{INT}!$G$2:$G${LASTROW}"
    Dc = f"{INT}!$D$2:$D${LASTROW}"
    Ec = f"{INT}!$E$2:$E${LASTROW}"
    QCOL = {"BOD": f"{INT}!$H$2:$H${LASTROW}", "TOC": f"{INT}!$I$2:$I${LASTROW}",
            "SS": f"{INT}!$J$2:$J${LASTROW}", "TN": f"{INT}!$K$2:$K${LASTROW}",
            "TP": f"{INT}!$L$2:$L${LASTROW}"}

    first_L, last_L = "B", get_column_letter(1 + len(years))

    # ---------------- 1_연도별_유입하수량 ----------------
    ws1 = wb.create_sheet("1_연도별_유입하수량")
    ncol = 1 + len(years) + 1
    set_title(ws1, 1, f"[표 1] {cfg['facility_name']} 연도별 유입하수량(전체)", ncol, "(단위: ㎥/일, %, 일)")
    hdr_row = 3
    ws1.cell(row=hdr_row, column=1, value="구  분")
    for i, y in enumerate(years): ws1.cell(row=hdr_row, column=2 + i, value=f"{y}년")
    avg_col = 2 + len(years)
    ws1.cell(row=hdr_row, column=avg_col, value="평균")
    style_row(ws1, hdr_row, ncol, header=True)

    r = hdr_row + 1
    row_avg = r
    ws1.cell(row=r, column=1, value="일평균(㎥/일)")
    for i, y in enumerate(years): ws1.cell(row=r, column=2 + i, value=f"=AVERAGEIFS({Fc},{Bc},{y})")
    ws1.cell(row=r, column=avg_col, value=f"=AVERAGE({first_L}{r}:{last_L}{r})")
    style_row(ws1, r, ncol, avg_col=avg_col)
    for c in range(2, ncol + 1): ws1.cell(row=r, column=c).number_format = "#,##0"
    
    r += 1
    ws1.cell(row=r, column=1, value="일평균/시설용량(%)")
    for i, y in enumerate(years):
        col = get_column_letter(2 + i)
        ws1.cell(row=r, column=2 + i, value=f"={col}{row_avg}/{CAP}")
    ws1.cell(row=r, column=avg_col, value=f"=AVERAGE({first_L}{r}:{last_L}{r})")
    style_row(ws1, r, ncol, avg_col=avg_col)
    for c in range(2, ncol + 1): ws1.cell(row=r, column=c).number_format = "0.0%"
    
    r += 1
    ws1.cell(row=r, column=1, value="일최대(㎥/일)")
    for i, y in enumerate(years): ws1.cell(row=r, column=2 + i, value=f"=_xlfn.MAXIFS({Fc},{Bc},{y})")
    ws1.cell(row=r, column=avg_col, value=f"=AVERAGE({first_L}{r}:{last_L}{r})")
    style_row(ws1, r, ncol, avg_col=avg_col)
    for c in range(2, ncol + 1): ws1.cell(row=r, column=c).number_format = "#,##0"
    
    r += 1
    ws1.cell(row=r, column=1, value="시설용량 초과일수(일)")
    for i, y in enumerate(years): ws1.cell(row=r, column=2 + i, value=f'=COUNTIFS({Bc},{y},{Fc},">"&{CAP})')
    ws1.cell(row=r, column=avg_col, value=f"=AVERAGE({first_L}{r}:{last_L}{r})")
    style_row(ws1, r, ncol, avg_col=avg_col)
    for c in range(2, ncol + 1): ws1.cell(row=r, column=c).number_format = "0.0"
    
    note(ws1, r + 2, ncol, "※ 유입하수량은 실제유입량(반류수 제외) 기준. 시설용량은 '0_안내' 시트 입력값을 참조.")
    ws1.column_dimensions["A"].width = 22
    for i in range(len(years) + 1): ws1.column_dimensions[get_column_letter(2 + i)].width = 12

    # ---------------- 2_청천강우_유입하수량 ----------------
    ws2 = wb.create_sheet("2_청천강우_유입하수량")
    set_title(ws2, 1, f"[표 2] {cfg['facility_name']} 연도별·강우사상별 유입하수량", ncol, "(단위: ㎥/일, 일)")
    blocks = [("청  천  시", "청천시"), ("강 우 영 향 일", "강우영향일"), ("강  우  시", "강우시")]
    r = 3
    for title, cat in blocks:
        ws2.cell(row=r, column=1, value=title)
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        style_row(ws2, r, ncol, block=True)
        r += 1
        hdr_r = r
        ws2.cell(row=hdr_r, column=1, value="구  분")
        for i, y in enumerate(years): ws2.cell(row=hdr_r, column=2 + i, value=f"{y}년")
        ws2.cell(row=hdr_r, column=avg_col, value="평균")
        style_row(ws2, hdr_r, ncol, header=True)
        r += 1
        ws2.cell(row=r, column=1, value="유입하수량 평균(㎥/일)")
        for i, y in enumerate(years): ws2.cell(row=r, column=2 + i, value=f'=AVERAGEIFS({Fc},{Bc},{y},{Gc},"{cat}")')
        ws2.cell(row=r, column=avg_col, value=f"=AVERAGE({first_L}{r}:{last_L}{r})")
        style_row(ws2, r, ncol, avg_col=avg_col)
        for c in range(2, ncol + 1): ws2.cell(row=r, column=c).number_format = "#,##0"
        r += 1
        ws2.cell(row=r, column=1, value="유입하수량 최대(㎥/일)")
        for i, y in enumerate(years): ws2.cell(row=r, column=2 + i, value=f'=IFERROR(_xlfn.MAXIFS({Fc},{Bc},{y},{Gc},"{cat}"),0)')
        ws2.cell(row=r, column=avg_col, value=f"=AVERAGE({first_L}{r}:{last_L}{r})")
        style_row(ws2, r, ncol, avg_col=avg_col)
        for c in range(2, ncol + 1): ws2.cell(row=r, column=c).number_format = "#,##0"
        r += 1
        ws2.cell(row=r, column=1, value="시설용량 초과일수(일)")
        for i, y in enumerate(years): ws2.cell(row=r, column=2 + i, value=f'=COUNTIFS({Bc},{y},{Gc},"{cat}",{Fc},">"&{CAP})')
        ws2.cell(row=r, column=avg_col, value=f"=AVERAGE({first_L}{r}:{last_L}{r})")
        style_row(ws2, r, ncol, avg_col=avg_col)
        for c in range(2, ncol + 1): ws2.cell(row=r, column=c).number_format = "0.0"
        r += 2
    note(ws2, r, ncol, "※ 강우시/강우영향일/청천시 정의는 '0_안내' 시트 참조.")
    ws2.column_dimensions["A"].width = 22
    for i in range(len(years) + 1): ws2.column_dimensions[get_column_letter(2 + i)].width = 12

    # ---------------- 3_강우별_유입하수량 ----------------
    ws3 = wb.create_sheet("3_강우별_유입하수량")
    bands = cfg["rain_bands"]
    band_defs = [("청천시", f"<{bands[0]}")] + [(f"강우시 {b}mm 이상", f">={b}") for b in bands]
    ncol3 = 1 + len(band_defs) + 1
    set_title(ws3, 1, f"[표 3] {cfg['facility_name']} 강우량별 유입하수량", ncol3, "(단위: ㎥/일)")
    hdr_row1 = 3
    ws3.cell(row=hdr_row1, column=1, value="구  분")
    ws3.merge_cells(start_row=hdr_row1, start_column=1, end_row=hdr_row1 + 1, end_column=1)
    ws3.cell(row=hdr_row1, column=2, value="청 천 시")
    ws3.merge_cells(start_row=hdr_row1, start_column=2, end_row=hdr_row1 + 1, end_column=2)
    ws3.cell(row=hdr_row1, column=3, value="강 우 시")
    ws3.merge_cells(start_row=hdr_row1, start_column=3, end_row=hdr_row1, end_column=2 + len(bands))
    ws3.cell(row=hdr_row1, column=ncol3, value="비고")
    ws3.merge_cells(start_row=hdr_row1, start_column=ncol3, end_row=hdr_row1 + 1, end_column=ncol3)
    style_row(ws3, hdr_row1, ncol3, header=True)
    hdr_row2 = hdr_row1 + 1
    for i, b in enumerate(bands): ws3.cell(row=hdr_row2, column=3 + i, value=f"{b}mm 이상")
    style_row(ws3, hdr_row2, ncol3, header=True)
    r = hdr_row2 + 1
    first_r3 = r
    for y in years:
        ws3.cell(row=r, column=1, value=f"{y}년")
        for i, (label, crit) in enumerate(band_defs):
            ws3.cell(row=r, column=2 + i, value=f'=AVERAGEIFS({Fc},{Bc},{y},{Ec},"{crit}")')
            ws3.cell(row=r, column=2 + i).number_format = "#,##0"
        style_row(ws3, r, ncol3)
        r += 1
    last_r3 = r - 1
    ws3.cell(row=r, column=1, value="평  균")
    for i in range(len(band_defs)):
        col = get_column_letter(2 + i)
        ws3.cell(row=r, column=2 + i, value=f"=AVERAGE({col}{first_r3}:{col}{last_r3})")
        ws3.cell(row=r, column=2 + i).number_format = "#,##0"
    for c in range(1, ncol3 + 1):
        ws3.cell(row=r, column=c).font = Font(name=FONT, bold=True, size=9)
        ws3.cell(row=r, column=c).fill = AVG_FILL
        ws3.cell(row=r, column=c).border = BORDER
    note(ws3, r + 2, ncol3, f"※ 청천시/강우시 구간은 당일 강수량 기준 누적(포함) 구간입니다. '0_안내' 시트 참조.")
    ws3.column_dimensions["A"].width = 14
    for i in range(2, ncol3 + 1): ws3.column_dimensions[get_column_letter(i)].width = 14

    # ---------------- 4_계절별_유입하수량 ----------------
    ws4 = wb.create_sheet("4_계절별_유입하수량")
    cols4 = [("전기간", None), ("우천일(강우시)", "강우시"), ("강우영향일", "강우영향일"), ("청천일", "청천시")]
    ncol4 = 1 + len(cols4) + 1
    set_title(ws4, 1, f"[표 4] {cfg['facility_name']} 계절별 유입하수량 분석", ncol4, "(단위: ㎥/일, %)")
    hdr_row4 = 3
    ws4.cell(row=hdr_row4, column=1, value="구  분")
    for i, (label, _) in enumerate(cols4): ws4.cell(row=hdr_row4, column=2 + i, value=label)
    ws4.cell(row=hdr_row4, column=2 + len(cols4), value="비고")
    style_row(ws4, hdr_row4, ncol4, header=True)
    seasons = ["봄", "여름", "가을", "겨울"]
    season_inflow_row = {}
    r = hdr_row4 + 1
    for season in seasons:
        inflow_row, ratio_row = r, r + 1
        season_inflow_row[season] = inflow_row
        ws4.cell(row=inflow_row, column=1, value=season)
        ws4.merge_cells(start_row=inflow_row, start_column=1, end_row=ratio_row, end_column=1)
        for i, (label, cat) in enumerate(cols4):
            col = get_column_letter(2 + i)
            f = f'=AVERAGEIFS({Fc},{Dc},"{season}")' if cat is None else f'=AVERAGEIFS({Fc},{Dc},"{season}",{Gc},"{cat}")'
            ws4.cell(row=inflow_row, column=2 + i, value=f)
            ws4.cell(row=inflow_row, column=2 + i).number_format = "#,##0"
            ws4.cell(row=ratio_row, column=2 + i, value=f"={col}{inflow_row}/{CAP}")
            ws4.cell(row=ratio_row, column=2 + i).number_format = "0.0%"
        ws4.cell(row=inflow_row, column=2 + len(cols4), value="유입량(㎥/일)")
        ws4.cell(row=ratio_row, column=2 + len(cols4), value="시설용량 대비 비율")
        for rr in (inflow_row, ratio_row):
            for c in range(1, ncol4 + 1):
                cell = ws4.cell(row=rr, column=c)
                cell.border = BORDER
                cell.font = BASE_FONT if c > 1 else Font(name=FONT, bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        r += 2
    note(ws4, r + 1, ncol4, "※ 우천일/강우영향일/청천일 정의는 '0_안내' 시트 참조. 비율은 시설용량 대비임.")

    hdrs4 = ["구  분", "시설용량", "전기간", "우천일(강우시)", "강우영향일", "청천일"]
    for i, h in enumerate(hdrs4):
        c = ws4.cell(row=3, column=8 + i, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    rr = 4
    for season in seasons:
        inflow_row = season_inflow_row[season]
        ws4.cell(row=rr, column=8, value=season)
        ws4.cell(row=rr, column=9, value=f"={CAP}")
        ws4.cell(row=rr, column=10, value=f"=B{inflow_row}")
        ws4.cell(row=rr, column=11, value=f"=C{inflow_row}")
        ws4.cell(row=rr, column=12, value=f"=D{inflow_row}")
        ws4.cell(row=rr, column=13, value=f"=E{inflow_row}")
        for c in range(8, 14):
            cell = ws4.cell(row=rr, column=c)
            cell.font = BASE_FONT; cell.border = BORDER
            cell.number_format = "#,##0"; cell.alignment = Alignment(horizontal="center")
        rr += 1
    ws4.cell(row=1, column=8, value="[차트용 데이터]").font = NOTE_FONT
    ws4.column_dimensions["A"].width = 10
    for i in range(2, ncol4 + 1): ws4.column_dimensions[get_column_letter(i)].width = 14
    for c, w in zip(range(8, 14), [10, 11, 11, 13, 12, 11]): ws4.column_dimensions[get_column_letter(c)].width = w

    # ---------------- 5_유입수질 ----------------
    ws5 = wb.create_sheet("5_유입수질")
    QITEMS = ["BOD", "TOC", "SS", "TN", "TP"]
    QLABEL = {"BOD": "BOD", "TOC": "TOC", "SS": "SS", "TN": "T-N", "TP": "T-P"}
    ncol5 = 1 + len(QITEMS) + 1
    set_title(ws5, 1, f"[표 5] {cfg['facility_name']} 유입수질 분석", ncol5, "(단위: ㎎/L, %)")

    r = 3
    ws5.cell(row=r, column=1, value="[1] 전기간 · 강우시 · 비강우시 유입수질")
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol5)
    style_row(ws5, r, ncol5, block=True)
    r += 1
    hdr_r = r
    ws5.cell(row=hdr_r, column=1, value="구  분")
    for i, q in enumerate(QITEMS): ws5.cell(row=hdr_r, column=2 + i, value=QLABEL[q])
    ws5.cell(row=hdr_r, column=2 + len(QITEMS), value="비고")
    style_row(ws5, hdr_r, ncol5, header=True)
    r += 1
    design_row = r
    ws5.cell(row=r, column=1, value="설계 유입수질")
    for i, q in enumerate(QITEMS):
        ws5.cell(row=r, column=2 + i, value=f'=IF({DES[q]}="","",{DES[q]})')
        ws5.cell(row=r, column=2 + i).number_format = "#,##0.0"
    style_row(ws5, r, ncol5)
    r += 1

    def qa_block(r0, label, formula_fn):
        rr = r0
        ws5.cell(row=rr, column=1, value=f"{label} 평균수질")
        for i, q in enumerate(QITEMS):
            ws5.cell(row=rr, column=2 + i, value=formula_fn(QCOL[q]))
            ws5.cell(row=rr, column=2 + i).number_format = "#,##0.0"
        style_row(ws5, rr, ncol5)
        rr += 1
        ws5.cell(row=rr, column=1, value=f"{label} 비율(%, 설계대비)")
        for i, q in enumerate(QITEMS):
            col = get_column_letter(2 + i)
            ws5.cell(row=rr, column=2 + i, value=f'=IFERROR({col}{rr-1}/{col}{design_row},"")')
            ws5.cell(row=rr, column=2 + i).number_format = "0.0%"
        style_row(ws5, rr, ncol5)
        rr += 1
        return rr

    r = qa_block(r, "전기간", lambda rng: f"=AVERAGE({rng})")
    r = qa_block(r, "강우시", lambda rng: f'=AVERAGEIFS({rng},{Gc},"강우시")')
    r = qa_block(r, "비강우시", lambda rng: f'=AVERAGEIFS({rng},{Gc},"<>강우시")')
    note(ws5, r, ncol5, "※ 강우시: 당일 강수량≥기준치 / 비강우시: 청천시+강우영향일. 설계수질 미입력 시 비율은 공란.")
    r += 2

    ws5.cell(row=r, column=1, value="[2] 연도별 평균 유입수질 및 설계기준 대비 비율")
    ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol5)
    style_row(ws5, r, ncol5, block=True)
    r += 1
    hdr_r2 = r
    ws5.cell(row=hdr_r2, column=1, value="구  분")
    for i, q in enumerate(QITEMS): ws5.cell(row=hdr_r2, column=2 + i, value=QLABEL[q])
    ws5.cell(row=hdr_r2, column=2 + len(QITEMS), value="비고")
    style_row(ws5, hdr_r2, ncol5, header=True)
    r += 1
    design_row2 = r
    ws5.cell(row=r, column=1, value="설계 유입수질")
    for i, q in enumerate(QITEMS):
        ws5.cell(row=r, column=2 + i, value=f'=IF({DES[q]}="","",{DES[q]})')
        ws5.cell(row=r, column=2 + i).number_format = "#,##0.0"
    style_row(ws5, r, ncol5)
    r += 1
    for y in years:
        ws5.cell(row=r, column=1, value=f"{y}년 평균수질")
        for i, q in enumerate(QITEMS):
            ws5.cell(row=r, column=2 + i, value=f"=AVERAGEIFS({QCOL[q]},{Bc},{y})")
            ws5.cell(row=r, column=2 + i).number_format = "#,##0.0"
        style_row(ws5, r, ncol5)
        r += 1
        ws5.cell(row=r, column=1, value=f"{y}년 비율(%, 설계대비)")
        for i, q in enumerate(QITEMS):
            col = get_column_letter(2 + i)
            ws5.cell(row=r, column=2 + i, value=f'=IFERROR({col}{r-1}/{col}{design_row2},"")')
            ws5.cell(row=r, column=2 + i).number_format = "0.0%"
        style_row(ws5, r, ncol5)
        r += 1

    def multi_year_block(r0, label, y0, y1):
        rr = r0
        ws5.cell(row=rr, column=1, value=f"{label} 평균수질")
        for i, q in enumerate(QITEMS):
            ws5.cell(row=rr, column=2 + i, value=f'=AVERAGEIFS({QCOL[q]},{Bc},">="&{y0},{Bc},"<="&{y1})')
            ws5.cell(row=rr, column=2 + i).number_format = "#,##0.0"
        style_row(ws5, rr, ncol5)
        for c in range(1, ncol5 + 1): ws5.cell(row=rr, column=c).fill = AVG_FILL
        rr += 1
        ws5.cell(row=rr, column=1, value=f"{label} 비율(%, 설계대비)")
        for i, q in enumerate(QITEMS):
            col = get_column_letter(2 + i)
            ws5.cell(row=rr, column=2 + i, value=f'=IFERROR({col}{rr-1}/{col}{design_row2},"")')
            ws5.cell(row=rr, column=2 + i).number_format = "0.0%"
        style_row(ws5, rr, ncol5)
        for c in range(1, ncol5 + 1): ws5.cell(row=rr, column=c).fill = AVG_FILL
        rr += 1
        return rr

    if len(years) >= 3:
        r = multi_year_block(r, f"과거 3년({years[-3]}~{years[-1]}) 평균", years[-3], years[-1])
    r = multi_year_block(r, f"과거 {len(years)}년({years[0]}~{years[-1]}) 평균", years[0], years[-1])

    toc_note = ""
    if any(toc_converted_flags):
        toc_note = f" ※ TOC 미측정 연도는 TOC=COD×{cfg['toc_cod_factor']} 환산값을 적용함('0_안내' 시트 참조)."
    note(ws5, r + 1, ncol5, f"※ 설계 유입수질은 '0_안내' 시트에서 입력합니다.{toc_note}")
    ws5.column_dimensions["A"].width = 26
    for i in range(2, ncol5 + 1): ws5.column_dimensions[get_column_letter(i)].width = 13

    chart5_anchor_row = r + 3

    # ---------------- 아름다운 디자인의 차트 생성 ----------------
    cats = Reference(ws_i, min_col=1, min_row=2, max_row=LASTROW)

    year_range = f"({years[0]}~{years[-1]})" if len(years) > 1 else f"({years[0]})"

    def style_series(s, color, width=6480, dash=None, no_line=False, marker=None, marker_size=4):
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = width
        s.graphicalProperties.line.cap = "rnd"
        s.graphicalProperties.solidFill = color
        if dash: s.graphicalProperties.line.dashStyle = dash
        if no_line: s.graphicalProperties.line.noFill = True
        if marker:
            s.marker.symbol = marker
            s.marker.size = marker_size
            s.marker.graphicalProperties.solidFill = color
            s.marker.graphicalProperties.line.solidFill = color
        else:
            s.marker.symbol = "none"
        s.smooth = False

    # 차트1 (유입하수량 + 시설용량 + 강수량)
    line1 = LineChart()
    v_in = Reference(ws_i, min_col=6, min_row=1, max_row=LASTROW)
    v_cap = Reference(ws_i, min_col=13, min_row=1, max_row=LASTROW)
    line1.add_data(v_in, titles_from_data=True)
    line1.add_data(v_cap, titles_from_data=True)
    line1.set_categories(cats)
    style_series(line1.series[0], "404040")
    style_series(line1.series[1], "C00000", width=12700, dash="dash")
    line1.x_axis.number_format = "yy-mm"
    line1.x_axis.majorTimeUnit = "months"
    line1.x_axis.tickLblSkip = 6
    line1.height = 9; line1.width = 26

    bar1 = BarChart()
    v_rain = Reference(ws_i, min_col=5, min_row=1, max_row=LASTROW)
    bar1.add_data(v_rain, titles_from_data=True)
    bar1.set_categories(cats)
    bar1.series[0].graphicalProperties.solidFill = "9DC3E6"
    bar1.series[0].graphicalProperties.line.noFill = True
    
    # ---------------------------------------------
    # (추가) 강수량 데이터 값 레이블 표현 적용
    bar1.dLbls = DataLabelList()
    bar1.dLbls.showVal = True
    # ---------------------------------------------
    
    bar1.y_axis.axId = 200
    bar1.y_axis.scaling.orientation = "maxMin"
    bar1.y_axis.crosses = "max"
    line1.y_axis.crosses = "autoZero"
    line1 += bar1

    apply_beautiful_chart_style(line1, title_text=f"일별 유입하수량 및 시설용량·강수량 추이{year_range}")
    style_axis(line1.y_axis, title_text="유입하수량(㎥/일)", show_gridlines=True, show_minor_gridlines=True)
    style_axis(bar1.y_axis, title_text="강수량(mm)", number_format="#,##0.00")
    ws1.add_chart(line1, "A13")

    # 차트4 (계절별 막대)
    bar4 = BarChart()
    bar4.type = "col"
    bar4.grouping = "clustered"
    cats4 = Reference(ws4, min_col=8, min_row=4, max_row=7)
    data4 = Reference(ws4, min_col=9, max_col=13, min_row=3, max_row=7)
    bar4.add_data(data4, titles_from_data=True)
    bar4.set_categories(cats4)
    for s, col in zip(bar4.series, ["C00000", "1F4E78", "2E75B6", "9DC3E6", "D9D9D9"]):
        s.graphicalProperties.solidFill = col
        s.graphicalProperties.line.noFill = True
    bar4.height = 9; bar4.width = 20
    apply_beautiful_chart_style(bar4, title_text=f"계절별 유입하수량 분석(전기간/우천일/강우영향일/청천일){year_range}")
    style_axis(bar4.y_axis, title_text="유입하수량(㎥/일)", show_gridlines=True, show_minor_gridlines=True)
    ws4.add_chart(bar4, "A" + str(rr + 2))

    # 차트5 (유입수질)
    line5 = LineChart()
    v_bod = Reference(ws_i, min_col=8, min_row=1, max_row=LASTROW)
    v_toc = Reference(ws_i, min_col=9, min_row=1, max_row=LASTROW)
    v_tn = Reference(ws_i, min_col=11, min_row=1, max_row=LASTROW)
    line5.add_data(v_bod, titles_from_data=True)
    line5.add_data(v_toc, titles_from_data=True)
    line5.add_data(v_tn, titles_from_data=True)
    line5.set_categories(cats)
    style_series(line5.series[0], "C00000")
    style_series(line5.series[1], "2E75B6")
    style_series(line5.series[2], "70AD47")
    line5.x_axis.number_format = "yyyy-mm-dd"
    line5.x_axis.majorTimeUnit = "months"
    line5.x_axis.tickLblSkip = 6
    line5.height = 10; line5.width = 26
    apply_beautiful_chart_style(line5, title_text=f"일별 유입수질(BOD·TOC·T-N) 변화 추이{year_range}")
    style_axis(line5.y_axis, title_text="유입수질(㎎/L)", show_gridlines=True, show_minor_gridlines=True)
    ws5.add_chart(line5, f"A{chart5_anchor_row}")

    # ---------------- 원본 데이터 시트 백업 ----------------
    if cfg["include_raw_sheets"]:
        for name, src_wb in zip(file_names, sewage_raw_wbs):
            sws = src_wb.active
            year_match = re.search(r"(20\d\d)", name)
            sheet_name = f"원본_{year_match.group(1)}" if year_match else f"원본_{sws.title}"[:31]
            wsr = wb.create_sheet(sheet_name[:31])
            for r_idx in range(1, sws.max_row + 1):
                for c_idx in range(1, sws.max_column + 1):
                    v = sws.cell(row=r_idx, column=c_idx).value
                    if v is not None:
                        wsr.cell(row=r_idx, column=c_idx, value=v)
            for c_idx in range(1, min(sws.max_column, 20) + 1):
                wsr.column_dimensions[get_column_letter(c_idx)].width = 10

        wsr = wb.create_sheet("원본_강수량")
        for i, h in enumerate(["날짜", "강수량(mm)"], start=1):
            wsr.cell(row=1, column=i, value=h)
        for i, rec in enumerate(rain_raw.itertuples(index=False), start=2):
            wsr.cell(row=i, column=1, value=rec.날짜).number_format = "yyyy-mm-dd"
            wsr.cell(row=i, column=2, value=float(rec.강수량))

    # 시트 순서 정리
    order = ["0_안내", "1_연도별_유입하수량", "2_청천강우_유입하수량", "3_강우별_유입하수량",
             "4_계절별_유입하수량", "5_유입수질", "일별통합데이터"]
    order += [s for s in wb.sheetnames if s not in order]
    wb._sheets = [wb[s] for s in order if s in wb.sheetnames]
    wb.active = 0
    return wb

# =====================================================================
# 2. Streamlit UI 구성
# =====================================================================
st.set_page_config(page_title="하수처리시설 분석기", layout="wide")
st.title("💧 하수처리시설 유입특성 분석 자동화")
st.markdown("운영현황 원본 엑셀과 기상청 강수량 파일을 업로드하면 분석 결과를 엑셀 파일로 생성해 줍니다.")

with st.sidebar:
    st.header("⚙️ 분석 설정")
    facility_name = st.text_input("처리시설명", "충청북도 단양군 매포 공공하수처리시설")
    capacity = st.number_input("시설용량 (㎥/일)", value=4000)
    
    st.subheader("설계 유입수질 (㎎/L)")
    design_bod = st.number_input("BOD", value=0.0)
    design_toc = st.number_input("TOC", value=0.0)
    design_ss = st.number_input("SS", value=0.0)
    design_tn = st.number_input("T-N", value=0.0)
    design_tp = st.number_input("T-P", value=0.0)

    st.subheader("분석 기준 설정")
    rain_event_mm = st.number_input("강우시 판정 기준 (mm)", value=3)
    rain_influence_days = st.number_input("강우영향일 일수", value=2)
    toc_cod_factor = st.number_input("TOC/COD 환산계수", value=0.625)

st.header("📂 데이터 업로드")
sewage_files = st.file_uploader("1️⃣ 연도별 운영현황 엑셀 업로드 (여러 파일 선택 가능)", type=["xlsx"], accept_multiple_files=True)
rainfall_file = st.file_uploader("2️⃣ 일별 강수량 엑셀 업로드", type=["xls", "xlsx", "csv"])

if st.button("🚀 분석 실행 및 엑셀 생성", type="primary"):
    if not sewage_files:
        st.error("운영현황 엑셀 파일을 최소 1개 이상 업로드해주세요.")
    elif not rainfall_file:
        st.error("강수량 엑셀 파일을 업로드해주세요.")
    else:
        cfg = {
            "facility_name": facility_name,
            "capacity": capacity,
            "design_quality": {
                "BOD": design_bod if design_bod > 0 else None,
                "TOC": design_toc if design_toc > 0 else None,
                "SS": design_ss if design_ss > 0 else None,
                "TN": design_tn if design_tn > 0 else None,
                "TP": design_tp if design_tp > 0 else None,
            },
            "rain_event_mm": rain_event_mm,
            "rain_influence_days": rain_influence_days,
            "rain_bands": [3, 6, 9], 
            "toc_cod_factor": toc_cod_factor,
            "include_raw_sheets": True
        }

        with st.spinner("데이터를 분석하고 예쁜 차트가 포함된 엑셀 파일을 생성 중입니다..."):
            try:
                sewage_dfs, sewage_raw_wbs, file_names, toc_flags = [], [], [], []
                for f in sewage_files:
                    df, raw_wb = extract_sewage_dataframe(f, toc_cod_factor=cfg["toc_cod_factor"])
                    sewage_dfs.append(df)
                    sewage_raw_wbs.append(raw_wb)
                    file_names.append(f.name)
                    toc_flags.append(df.attrs.get("toc_converted", False))

                rain_df, rain_raw = extract_rainfall_dataframe(rainfall_file)
                master_df = build_master(sewage_dfs, rain_df, cfg)
                
                wb = build_workbook(master_df, sewage_raw_wbs, rain_raw, cfg, file_names, toc_flags)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success("✅ 분석 완료! 아래 버튼을 눌러 결과 파일을 다운로드하세요.")
                st.download_button(
                    label="📥 결과 엑셀 파일 다운로드",
                    data=output,
                    file_name=f"{facility_name}_유입특성분석.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"오류가 발생했습니다. 파일 데이터를 처리할 수 없습니다: {e}")
